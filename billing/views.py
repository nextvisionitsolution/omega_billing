from django.shortcuts import render
from django.http import JsonResponse
import json
from django.core.serializers.json import DjangoJSONEncoder
from .models import Item, Bill, BillItem, Category
from django.db.models import Sum, F, DecimalField, IntegerField
from django.db.models import ExpressionWrapper
from django.db.models.functions import Coalesce
import openpyxl
from django.http import HttpResponse
from django.core.paginator import Paginator
from datetime import datetime, time, timedelta
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from decimal import Decimal, InvalidOperation
from .models import Shift
from .models import EditedBillHistory
from django.conf import settings
from django.utils.timezone import localtime
from openpyxl.styles import Font
import tempfile
import os
import pyzipper
from django.http import FileResponse
from django.contrib import messages
def pos_page(request):

    current_shift = get_current_shift()

    items = Item.objects.select_related(
        'category'
    ).all()

    item_data = []

    for i in items:

        sold_qty = BillItem.objects.filter(
            item=i
        ).aggregate(
            total=Sum('qty')
        )['total'] or 0

        item_data.append({
            "id": i.id,
            "name": i.name,
            "category": i.category.name if i.category else "",

            "price1": float(i.price1 or 0),
            "price2": float(i.price2 or 0),
            "price3": float(i.price3 or 0),

            "sold_qty": sold_qty
        })

    return render(
        request,
        "pos.html",
        {
            "items_json": json.dumps(
                item_data,
                cls=DjangoJSONEncoder
            ),

            "categories": Category.objects.all(),

            "current_shift": current_shift
        }
    )


def save_bill(request):

    if request.method == "POST":

        data = json.loads(request.body)

        current_shift = get_current_shift()

        bill = Bill.objects.create(

            customer_name=data.get("customer", ""),

            total=Decimal(
                str(data.get("total", 0))
            ),

            shift=current_shift.active_shift

        )

        for i in data.get("items", []):

            BillItem.objects.create(

                bill=bill,

                item_id=i["id"],

                price=Decimal(
                    str(i["price"])
                ),

                qty=i["qty"]

            )

        return JsonResponse({
            "status": "success"
        })


def sales_report_page(request):
    return render(request, "sales_report.html")

def sales_report(request):

    item_name = request.GET.get("item")

    shift = request.GET.get(
        "shift",
        "1"
    )

    page_number = request.GET.get(
        "page",
        1
    )

    bill_items = BillItem.objects.select_related(
        'bill',
        'item',
        'item__category'
    ).filter(

        # SHIFT FILTER
        bill__shift=shift

    )

    # SEARCH FILTER
    if item_name:

        bill_items = bill_items.filter(
            item__name__icontains=item_name
        )

    # EDITED ITEMS MAP
    edited_map = set(
        EditedBillHistory.objects.values_list(
            "bill_id",
            "item_id"
        )
    )

    rows = bill_items.annotate(

        line_total=ExpressionWrapper(

            F('qty') * F('price'),

            output_field=DecimalField(
                max_digits=15,
                decimal_places=2
            )

        )

    ).order_by('-bill__id')

    data = []

    edited_history = EditedBillHistory.objects.all()

    edited_map = set()

    for h in edited_history:
        edited_map.add(
            (h.bill_id, h.item_id)
        )

    for i in rows:

        data.append({

            "bill_no": i.bill.id,

            "date": i.bill.created_at.strftime(
                "%d-%m-%Y %I:%M %p"
            ),

            "item": i.item.name,

            "category": i.item.category.name if i.item.category else "",

            "qty": i.qty,

            "price": float(i.price),

            "amount": float(i.line_total or 0),

            "item_edited": (
                i.bill.id,
                i.item.id
            ) in edited_map,

            "edited_at": localtime(
                i.bill.edited_at
            ).strftime(
                "%d-%m-%Y %I:%M %p"
            ) if (
                (i.bill.id, i.item.id) in edited_map
                and i.bill.edited_at
            ) else ""

        })

    grand_total = sum(
        x["amount"] for x in data
    )

    # CATEGORY STATS

    category_stats = bill_items.values(
        'item__category__name'
    ).annotate(

        total_qty=Coalesce(

            Sum(
                'qty',
                output_field=IntegerField()
            ),

            0,

            output_field=IntegerField()

        )

    )

    cat_data = []

    for c in category_stats:

        cat_data.append({

            "category": c['item__category__name'] or "Other",

            "qty": c['total_qty']

        })

    paginator = Paginator(data, 10)

    page_obj = paginator.get_page(
        page_number
    )

    return JsonResponse({

        "items": list(
            page_obj.object_list
        ),

        "grand_total": grand_total,

        "current_page": page_obj.number,

        "total_pages": paginator.num_pages,

        "category_stats": cat_data

    })



def export_sales_excel(request):

    shift = request.GET.get("shift", "1")
    item_filter = request.GET.get("item", "").strip()

    # FILTER SALES
    sales = BillItem.objects.filter(

        bill__shift=shift

    ).select_related(

        "item",
        "item__category"

    )

    # ITEM FILTER
    if item_filter:

        sales = sales.filter(
            item__name__icontains=item_filter
        )

    # GROUPING
    sales = sales.values(

        "item__category__name",
        "item__name",
        "price"

    ).annotate(

        total_qty=Sum("qty"),

        total_amount=Sum(
            ExpressionWrapper(
                F("qty") * F("price"),
                output_field=DecimalField(
                    max_digits=15,
                    decimal_places=2
                )
            )
        )

    ).order_by(

        "item__category__name",
        "item__name"

    )

    # EXCEL
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = f"Shift {shift} Sales"

    # DATE row
    ws.append([
        f"Sales Report Date : {timezone.localdate().strftime('%d-%m-%Y')}"
    ])

    ws.append([])

    # HEADERS
    headers = [
        "Category",
        "Item",
        "Qty",
        "Rate",
        "Amount"
    ]

    ws.append(headers)

    grand_total = 0

    current_category = None
    current_category_total_qty = 0

    for s in sales:

        category = s["item__category__name"] or "Other"

        # category change zali tar previous category total print kar
        if current_category and current_category != category:

            row_no = ws.max_row + 1

            ws.append([
                "",
                "",
                f"{current_category} Total : {current_category_total_qty}",
                "",
                ""
            ])

            ws[f'C{row_no}'].font = Font(bold=True)

            ws.append([])

            current_category_total_qty = 0

        # CATEGORY TITLE
        if current_category != category:

            current_category = category

            ws.append([])
            ws.append([
                f"{category}"
            ])

        amount = float(s["total_amount"] or 0)

        grand_total += amount

        current_category_total_qty += s["total_qty"]

        ws.append([

            "",

            s["item__name"],

            s["total_qty"],

            float(s["price"]),

            amount

        ])

    # last category total
    if current_category:

        row_no = ws.max_row + 1

        ws.append([
            "",
            "",
            f"{current_category} Total : {current_category_total_qty}",
            "",
            ""
        ])

        ws[f'C{row_no}'].font = Font(bold=True)

    # GRAND TOTAL
    ws.append([])
    ws.append([
        "",
        "",
        "",
        "Grand Total",
        grand_total
    ])

    # AUTO WIDTH
    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = max_length + 4

        ws.column_dimensions[column_letter].width = adjusted_width

    # RESPONSE
    # TEMP EXCEL FILE
    temp_xlsx = tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        delete=False
    )

    wb.save(temp_xlsx.name)
    temp_xlsx.close()

    # PASSWORD PROTECTED ZIP
    zip_path = temp_xlsx.name.replace(
        ".xlsx",
        ".zip"
    )

    with pyzipper.AESZipFile(
        zip_path,
        "w",
        compression=pyzipper.ZIP_DEFLATED,
        encryption=pyzipper.WZ_AES
    ) as zf:

        zf.setpassword(
            settings.EXPORT_EXCEL_PASSWORD.encode()
        )

        zf.write(
            temp_xlsx.name,
            arcname=f"sales_shift_{shift}.xlsx"
        )

    # RETURN ZIP FILE
    response = FileResponse(
        open(zip_path, "rb"),
        as_attachment=True,
        filename=f"sales_shift_{shift}.zip"
    )

    return response



def delete_today_sales(request):
    today = timezone.now().date()

    # delete all bill items of today bills
    BillItem.objects.filter(bill__created_at__date=today).delete()

    # delete bills of today
    Bill.objects.filter(created_at__date=today).delete()

    return JsonResponse({"status": "success", "message": "Today's data deleted"})

def delete_all_sales(request):

    # delete all bill items first
    BillItem.objects.all().delete()

    # then delete all bills
    Bill.objects.all().delete()

    return JsonResponse({
        "status": "success",
        "message": "All sales data deleted"
    })





# ================= ITEM CRUD =================

def item_list(request):

    items=Item.objects.select_related(
    'category'
    ).all().order_by('-id')

    return render(
        request,
        'items/item_list.html',
        {
            'items': items
        }
    )


def add_item(request):

    if request.method == "POST":

        # EXCEL UPLOAD
        if request.FILES.get("excel_file"):

            try:
                excel_file = request.FILES["excel_file"]

                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active

                for row in ws.iter_rows(min_row=2, values_only=True):

                    category_name = row[0]
                    item_name = row[1]

                    # Convert safely
                    try:
                        price1 = Decimal(str(row[2] or 0).strip())
                        price2 = Decimal(str(row[3] or 0).strip())
                        price3 = Decimal(str(row[4] or 0).strip())
                    except (InvalidOperation, ValueError):
                        messages.error(
                            request,
                            f"Invalid price value found for item: {item_name}. Please upload correct Excel format."
                        )
                        return redirect("add_item")

                    category, created = Category.objects.get_or_create(
                        name=category_name
                    )

                    Item.objects.update_or_create(
                        name=item_name,
                        category=category,
                        defaults={
                            "price1": price1,
                            "price2": price2,
                            "price3": price3,
                        }
                    )

                messages.success(request, "Excel imported successfully.")
                return redirect("item_list")

            except Exception:
                messages.error(
                    request,
                    "Invalid Excel file. Please upload a valid Excel file with correct format."
                )
                return redirect("add_item")

        # SINGLE ITEM SAVE

        try:
            Item.objects.create(
                name=request.POST.get("name"),
                category_id=request.POST.get("category"),
                price1=Decimal(request.POST.get("price1") or 0),
                price2=Decimal(request.POST.get("price2") or 0),
                price3=Decimal(request.POST.get("price3") or 0),
            )

            messages.success(request, "Item added successfully.")

        except Exception:
            messages.error(request, "Please enter valid price values.")

        return redirect("add_item")

    return render(
        request,
        "items/add_item.html",
        {
            "categories": Category.objects.all()
        }
    )


def edit_item(request, pk):

    item=get_object_or_404(
        Item,
        pk=pk
    )

    if request.method == "POST":

        item.name = request.POST.get("name")
        item.category_id = request.POST.get("category")
        item.price1 = request.POST.get("price1")
        item.price2 = request.POST.get("price2")

        price3 = request.POST.get("price3")
        item.price3 = float(price3) if price3 else 0

        item.save()

        return redirect('item_list')

    return render(
        request,
        'items/edit_item.html',
        {
            'item': item,
            'categories': Category.objects.all()
        }
    )


def delete_item(request, pk):

    item = get_object_or_404(Item, pk=pk)
    item.delete()

    return redirect('item_list')

# ================= CATEGORY =================

def category_list(request):

    categories = Category.objects.all().order_by('-id')

    return render(
        request,
        'category/category_list.html',
        {
            'categories':categories
        }
    )


def add_category(request):

    if request.method=="POST":

        name=request.POST.get("name")

        Category.objects.create(
            name=name
        )

        return redirect(
            'category_list'
        )

    return render(
        request,
        'category/add_category.html'
    )


def edit_category(request,pk):

    category=get_object_or_404(
        Category,
        pk=pk
    )

    if request.method=="POST":

        category.name=request.POST.get(
            "name"
        )

        category.save()

        return redirect(
            'category_list'
        )

    return render(
        request,
        'category/edit_category.html',
        {
            'category':category
        }
    )


def delete_category(request,pk):

    category=get_object_or_404(
        Category,
        pk=pk
    )

    category.delete()

    return redirect(
        'category_list'
    )

def get_shift_range(date=None):

    if not date:
        date = timezone.localdate()

    current_tz = timezone.get_current_timezone()

    shift_start = timezone.make_aware(
        datetime.combine(
            date,
            time(13, 0)
        ),
        current_tz
    )

    shift_end = shift_start + timedelta(days=1)

    return shift_start, shift_end

def get_current_shift():

    shift = Shift.objects.filter(
        is_active=True
    ).first()

    if not shift:

        shift = Shift.objects.create(
            active_shift='1'
        )

    return shift

def close_shift(request):

    current_shift = get_current_shift()

    current_shift.is_active = False
    current_shift.closed_at = timezone.now()
    current_shift.save()

    next_shift = '2' if current_shift.active_shift == '1' else '1'

    Shift.objects.create(
        active_shift=next_shift
    )

    return JsonResponse({
        "status": "success",
        "next_shift": next_shift
    })

def edit_bill(request, bill_id):

    if not request.session.get(f"edit_bill_auth_{bill_id}"):
        return redirect(
            "verify_edit_bill",
            bill_id=bill_id
        )

    bill = get_object_or_404(
        Bill,
        id=bill_id
    )

    bill_items = BillItem.objects.select_related(
        'item'
    ).filter(
        bill=bill
    )

    if request.method == "POST":

        total = 0

        for item in bill_items:

            old_qty = item.qty
            old_price = item.price

            new_qty = int(
                request.POST.get(f"qty_{item.id}")
            )

            new_price = Decimal(
                request.POST.get(f"price_{item.id}")
            )

            # SAVE HISTORY ONLY IF CHANGED
            if (
                old_qty != new_qty or
                old_price != new_price
            ):

                EditedBillHistory.objects.create(

                    bill=bill,

                    item=item.item,

                    old_qty=old_qty,

                    new_qty=new_qty,

                    old_price=old_price,

                    new_price=new_price

                )

            item.qty = new_qty
            item.price = new_price

            item.save()

            total += item.qty * item.price

        bill.total = total

        bill.is_edited = True
        bill.edited_at = timezone.now()
        bill.save()

        request.session.pop(
            f"edit_bill_auth_{bill_id}",
            None
        )

        return redirect(
            "sales_report_page"
        )

    return render(
        request,
        "edit_bill.html",
        {
            "bill": bill,
            "bill_items": bill_items
        }
    )

def edited_bills(request):

    history = EditedBillHistory.objects.select_related(
        'bill',
        'item'
    ).order_by(
        '-edited_at'
    )

    data = []
    seen = set()

    for h in history:

        key = (
            h.bill_id,
            h.item_id
        )

        if key in seen:
            continue

        seen.add(key)

        data.append({

            "id": h.bill.id,

            "bill_id": h.bill.id,

            "item_name": h.item.name,

            "qty": h.new_qty,

            "price": h.new_price,

            "total": h.new_qty * h.new_price,

            "edited_at": h.edited_at

        })

    return render(
        request,
        "edited_bills.html",
        {
            "edited_items": data
        }
    )

def edited_bill_detail(request, id):

    bill = get_object_or_404(
        Bill,
        id=id
    )

    history = EditedBillHistory.objects.select_related(
        'item'
    ).filter(
        bill=bill
    ).order_by('-edited_at')

    return render(
        request,
        "edited_bill_detail.html",
        {
            "bill": bill,
            "history": history
        }
    )

def verify_edit_bill(request, bill_id):
    if request.method == "POST":
        password = request.POST.get("password")

        if password == settings.EDIT_BILL_PASSWORD:
            request.session[f"edit_bill_auth_{bill_id}"] = True

            return redirect(
                "edit_bill",
                bill_id=bill_id
            )
        
        return render(
            request,
            "verify_edit_bill.html",
            {
                "bill_id": bill_id,
                "error": "Invalid Password"
            }
        )
    
    return render(
        request,
        "verify_edit_bill.html",
        {
            "bill_id": bill_id
        }
    )

def bill_report(request):

    bills = Bill.objects.all().order_by("-id")

    paginator = Paginator(bills, 10)

    page_obj = paginator.get_page(
        request.GET.get("page")
    )

    return render(
        request,
        "bill_report.html",
        {
            "bills": page_obj,
            "page_obj": page_obj
        }
    )

def bill_detail(request, bill_id):

    bill = get_object_or_404(
        Bill,
        id=bill_id
    )

    items = BillItem.objects.filter(
        bill=bill
    ).select_related("item")

    items_data = []

    for i in items:

        items_data.append({

            "item": i.item.name,

            "qty": i.qty,

            "price": i.price,

            "amount": i.qty * i.price

        })

    return render(
        request,
        "bill_detail.html",
        {
            "bill": bill,
            "items": items_data
        }
    )

def delete_shift_sales(request):

    shift = request.GET.get("shift")

    if not shift:
        return JsonResponse({
            "status": "error",
            "message": "Shift required"
        })

    # त्या shift चे bill items delete
    BillItem.objects.filter(
        bill__shift=shift
    ).delete()

    # त्या shift चे bills delete
    Bill.objects.filter(
        shift=shift
    ).delete()

    return JsonResponse({
        "status": "success",
        "message": f"Shift {shift} data deleted"
    })